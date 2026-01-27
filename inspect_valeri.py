import openpyxl

wb = openpyxl.load_workbook('VALERI_V4.xlsx', data_only=True)
target_sheet = 'SEU-Maßnahmen-Liste'

if target_sheet in wb.sheetnames:
    print(f"Reading Sheet: {target_sheet}")
    ws = wb[target_sheet]
    # Print first 10 rows
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(x).strip() for x in row if x is not None]
        # filter empty strings
        vals = [v for v in vals if v]
        if vals:
            print(f"R{i}: {vals}")
else:
    print(f"{target_sheet} not found.")
